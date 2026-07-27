"""Phase-4 acceptance flows through the API: generate a draft from the
snapshot (known-good quantities), override, place (exports + gated email),
ingest the acceptance reply, confirm both proposals, watch the timeline; plus
domestic vendors, analogies, rules overrides, Canada seams, and RBAC.

Sales math used throughout (engine spec): steady 100 units/month, no stock,
no incoming, target 8 months -> sea = 8*100 = 800, air = 3*100 = 300 (the
3-month near-term floor); at 50/month -> 400 sea / 150 air.
"""

from __future__ import annotations

import io
from datetime import date

import openpyxl
from app.models import (
    OrderEventProposal,
    Product,
    Role,
    SalesMonthly,
    StockLevel,
    utcnow,
)
from app.ordering.export import EXPORT_COLUMNS
from sqlalchemy import select

from .util import login, mk_product, mk_user

LAMP_SKU = "HO0000500400"
DHOOP_SKU = "IN0001200500"

ACCEPTANCE_REPLY = (
    "Namaskaram, we can only send 200 of the 500 lamps, and dhoop sticks are "
    "discontinued. Everything else ships as planned."
)


def _seed_sales(db, product: Product, units_per_month: float, months: int = 12) -> None:
    """`months` COMPLETE months of steady sales ending last month."""
    today = utcnow().date()
    total = today.year * 12 + (today.month - 1)
    for back in range(1, months + 1):
        ordinal = total - back
        db.add(
            SalesMonthly(
                product_id=product.id,
                year=ordinal // 12,
                month=ordinal % 12 + 1,
                channel="pos",
                units=units_per_month,
            )
        )
    db.commit()


def _admin(db, client):
    mk_user(db, "admin@test.io", (Role.ADMIN, None, None))
    return login(client, "admin@test.io")


def _seed_import_products(db):
    lamp = mk_product(db, LAMP_SKU, "Brass Guru Puja Lamp Large", category="CONX", price=40)
    dhoop = mk_product(db, DHOOP_SKU, "Sambrani Dhoop Sticks", category="INCENSE", price=8)
    lamp.cost, dhoop.cost = 20, 3
    db.commit()
    _seed_sales(db, lamp, 100)
    _seed_sales(db, dhoop, 50)
    return lamp, dhoop


def test_full_import_order_lifecycle(db, client):
    headers = _admin(db, client)
    _seed_import_products(db)

    # recipients configured so the placement email has somewhere to go
    r = client.put(
        "/api/v1/ordering/email-settings",
        json={"india_to": ["orders@coimbatore.test"], "cc": ["office@ishalife.test"]},
        headers=headers,
    )
    assert r.status_code == 200

    # ---- generate the draft (known-good quantities from the seeded snapshot)
    r = client.post(
        "/api/v1/ordering/orders",
        json={"name": "Q3 2026", "destination": "III"},
        headers=headers,
    )
    assert r.status_code == 201, r.text
    detail = r.json()
    order_id = detail["order"]["id"]
    assert detail["order"]["status"] == "draft"
    lines = {ln["global_sku"]: ln for ln in detail["lines"]}
    assert lines[LAMP_SKU]["suggested_sea_qty"] == 800
    assert lines[LAMP_SKU]["suggested_air_qty"] == 300
    assert lines[DHOOP_SKU]["suggested_sea_qty"] == 400
    assert lines[DHOOP_SKU]["suggested_air_qty"] == 150
    # baseline shown alongside; flat history means they agree
    assert lines[LAMP_SKU]["baseline_sea_qty"] == 800
    assert lines[LAMP_SKU]["suggestion"]["air_split_reason"]
    assert lines[LAMP_SKU]["suggestion"]["projected_moh"] == [0.0] * 6

    # ---- buyer override on the draft: lamp sea 800 -> 500
    lamp_line = lines[LAMP_SKU]
    r = client.patch(
        f"/api/v1/ordering/orders/{order_id}/lines/{lamp_line['id']}",
        json={"final_sea_qty": 500},
        headers=headers,
    )
    assert r.status_code == 200
    assert r.json()["final_sea_qty"] == 500
    assert r.json()["suggested_sea_qty"] == 800  # the suggestion stays visible

    # ---- place: exports attached, legs created, email SIMULATED (flag off)
    r = client.post(f"/api/v1/ordering/orders/{order_id}/place", headers=headers)
    assert r.status_code == 200, r.text
    placed = r.json()
    assert placed["order"]["status"] == "placed"
    assert {leg["label"] for leg in placed["legs"]} == {"Q3 2026", "Q3 2026 AIR"}
    assert placed["email_gate_reason"]  # gate is ON (flag off) -> dry-run

    r = client.get(f"/api/v1/ordering/orders/{order_id}/timeline", headers=headers)
    timeline = r.json()
    assert [a["filename"] for a in timeline["attachments"]] == [
        "Q3 2026 ORDER LIST.csv",
        "Q3 2026 ORDER LIST.xlsx",
    ]
    outbound = [m for m in timeline["emails"] if m["direction"] == "out"]
    assert len(outbound) == 1
    assert outbound[0]["status"] == "simulated"  # rendered without sending
    assert "orders@coimbatore.test" in outbound[0]["recipients"]
    assert "Q3 2026" in outbound[0]["subject"]
    # the summary reflects the override: 500 (lamp, was 800) + 400 (dhoop) sea
    assert "Sea units:  900" in outbound[0]["body"]

    # placing twice is refused
    r = client.post(f"/api/v1/ordering/orders/{order_id}/place", headers=headers)
    assert r.status_code == 409

    # ---- the acceptance reply -> two correctly-parsed proposals with quotes
    r = client.post(
        f"/api/v1/ordering/orders/{order_id}/ingest-email",
        json={"sender": "coimbatore@ishalife.in", "subject": "Re: Q3 2026",
              "body": ACCEPTANCE_REPLY},
        headers=headers,
    )
    assert r.status_code == 201, r.text
    proposals = [p for p in r.json()["proposals"] if p["status"] == "pending"]
    assert len(proposals) == 2
    qty = next(p for p in proposals if p["kind"] == "qty_change")
    disc = next(p for p in proposals if p["kind"] == "discontinued")
    assert "200 of the 500 lamps" in qty["quote"]
    assert qty["line_sku"] == LAMP_SKU
    assert qty["payload"]["sea"] == {"from": 500, "to": 200}
    assert 0 < qty["confidence"] < 1
    assert "dhoop sticks are discontinued" in disc["quote"]
    assert disc["line_sku"] == DHOOP_SKU

    # ---- confirming both updates order state through the append-only log
    r = client.post(
        f"/api/v1/ordering/proposals/{qty['id']}/decide",
        json={"accept": True},
        headers=headers,
    )
    assert r.status_code == 200, r.text
    r = client.post(
        f"/api/v1/ordering/proposals/{disc['id']}/decide",
        json={"accept": True},
        headers=headers,
    )
    assert r.status_code == 200

    r = client.get(f"/api/v1/ordering/orders/{order_id}", headers=headers)
    lines = {ln["global_sku"]: ln for ln in r.json()["lines"]}
    assert lines[LAMP_SKU]["final_sea_qty"] == 200
    assert lines[LAMP_SKU]["origin_sea_qty"] == 800  # origin never moves
    assert lines[DHOOP_SKU]["line_status"] == "discontinued"
    assert lines[DHOOP_SKU]["final_sea_qty"] == 0

    r = client.get(f"/api/v1/ordering/orders/{order_id}/timeline", headers=headers)
    timeline = r.json()
    kinds = [e["kind"] for e in timeline["events"]]
    # origin -> override -> email out -> placed -> email in -> the two confirmations
    assert kinds.count("qty_change") == 2
    assert "discontinued" in kinds
    confirmed = [e for e in timeline["events"] if e["kind"] == "discontinued"][0]
    assert "dhoop sticks are discontinued" in confirmed["source_quote"]
    assert all(p["status"] != "pending" for p in timeline["proposals"])

    # a decided proposal can't be decided again
    r = client.post(
        f"/api/v1/ordering/proposals/{qty['id']}/decide",
        json={"accept": True},
        headers=headers,
    )
    assert r.status_code == 409


def test_export_matches_agreed_columns_and_opens_in_excel(db, client):
    headers = _admin(db, client)
    _seed_import_products(db)
    r = client.post(
        "/api/v1/ordering/orders", json={"name": "Q4 2026"}, headers=headers
    )
    order_id = r.json()["order"]["id"]

    r = client.get(f"/api/v1/ordering/orders/{order_id}/export.csv", headers=headers)
    assert r.status_code == 200
    header_line = r.text.splitlines()[0]
    assert header_line == ",".join(h for _, h in EXPORT_COLUMNS)
    assert "attachment" in r.headers["content-disposition"]

    r = client.get(f"/api/v1/ordering/orders/{order_id}/export.xlsx", headers=headers)
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active
    got_header = [ws.cell(1, c).value for c in range(1, len(EXPORT_COLUMNS) + 1)]
    assert got_header == [h for _, h in EXPORT_COLUMNS]
    # both seeded SKUs order something, so the sheet has 2 data rows
    skus = {ws.cell(r, 3).value for r in range(2, ws.max_row + 1)}
    assert skus == {LAMP_SKU, DHOOP_SKU}


def test_manual_events_split_and_attachment(db, client):
    headers = _admin(db, client)
    _seed_import_products(db)
    r = client.post(
        "/api/v1/ordering/orders", json={"name": "Q1 2027"}, headers=headers
    )
    detail = r.json()
    order_id = detail["order"]["id"]
    client.post(f"/api/v1/ordering/orders/{order_id}/place", headers=headers)

    # a split the parser missed, entered by hand -> a new leg
    r = client.post(
        f"/api/v1/ordering/orders/{order_id}/events",
        json={
            "kind": "split",
            "payload": {"label": "Q1 2027 ADD AIR", "method": "air",
                        "eta": str(date.today()), "lines": {LAMP_SKU: 60}},
            "note": "vendor splitting the balms onto a second air shipment",
        },
        headers=headers,
    )
    assert r.status_code == 201, r.text
    labels = {leg["label"] for leg in r.json()["legs"]}
    assert "Q1 2027 ADD AIR" in labels

    # method change moves quantity between legs on a line
    lamp_line = next(ln for ln in detail["lines"] if ln["global_sku"] == LAMP_SKU)
    r = client.post(
        f"/api/v1/ordering/orders/{order_id}/events",
        json={
            "kind": "method_change",
            "line_id": lamp_line["id"],
            "payload": {"from": "sea", "to": "air", "qty": 100},
        },
        headers=headers,
    )
    assert r.status_code == 201, r.text
    r = client.get(f"/api/v1/ordering/orders/{order_id}", headers=headers)
    lamp = next(ln for ln in r.json()["lines"] if ln["global_sku"] == LAMP_SKU)
    assert lamp["final_sea_qty"] == 700 and lamp["final_air_qty"] == 400

    # bad payloads are refused, state untouched
    r = client.post(
        f"/api/v1/ordering/orders/{order_id}/events",
        json={"kind": "qty_change", "line_id": lamp_line["id"], "payload": {}},
        headers=headers,
    )
    assert r.status_code == 400

    # attachment upload lands on the timeline
    r = client.post(
        f"/api/v1/ordering/orders/{order_id}/attachments",
        files={"file": ("customs.pdf", b"%PDF-1.4 fake", "application/pdf")},
        data={"note": "customs paperwork"},
        headers=headers,
    )
    assert r.status_code == 201
    attachment = next(
        a for a in r.json()["attachments"] if a["filename"] == "customs.pdf"
    )
    r = client.get(
        f"/api/v1/ordering/orders/{order_id}/attachments/{attachment['id']}/download",
        headers=headers,
    )
    assert r.status_code == 200 and r.content.startswith(b"%PDF")


def test_rejecting_a_proposal_changes_nothing(db, client):
    headers = _admin(db, client)
    _seed_import_products(db)
    r = client.post("/api/v1/ordering/orders", json={"name": "R1"}, headers=headers)
    order_id = r.json()["order"]["id"]
    client.post(f"/api/v1/ordering/orders/{order_id}/place", headers=headers)
    r = client.post(
        f"/api/v1/ordering/orders/{order_id}/ingest-email",
        json={"body": ACCEPTANCE_REPLY},
        headers=headers,
    )
    qty = next(p for p in r.json()["proposals"] if p["kind"] == "qty_change")
    r = client.post(
        f"/api/v1/ordering/proposals/{qty['id']}/decide",
        json={"accept": False, "note": "we'll push back on this"},
        headers=headers,
    )
    assert r.status_code == 200
    lines = {
        ln["global_sku"]: ln
        for ln in client.get(f"/api/v1/ordering/orders/{order_id}", headers=headers).json()["lines"]
    }
    assert lines[LAMP_SKU]["final_sea_qty"] == 800  # untouched
    proposal = db.execute(select(OrderEventProposal).where(OrderEventProposal.id == qty["id"])).scalar()
    assert proposal.status == "rejected"


def test_domestic_vendor_flow(db, client):
    headers = _admin(db, client)
    r = client.post(
        "/api/v1/ordering/vendors",
        json={"name": "Botanie Soap", "kind": "us", "contact_email": "orders@botanie.test"},
        headers=headers,
    )
    assert r.status_code == 201, r.text
    vendor_id = r.json()["id"]
    soap = mk_product(db, "BC0007100006", "Lavender & Comfrey Soap", category="BODY CARE", price=9)
    soap.vendor_id = vendor_id
    soap.moq = 1080
    db.commit()
    _seed_sales(db, soap, 164)
    db.add(StockLevel(product_id=soap.id, location_key="bwhse", qty=32))
    db.commit()

    # MOQ rule: MOH 32/164 ≈ 0.2 < 4 -> order one MOQ
    r = client.get(f"/api/v1/ordering/vendors/{vendor_id}/suggestions", headers=headers)
    assert r.status_code == 200
    item = next(i for i in r.json()["items"] if i["global_sku"] == soap.global_sku)
    assert item["suggested_sea_round"] == 1080
    assert "MOQ" in item["air_split_reason"]

    # a vendor-assigned product never appears on the India review
    r = client.post("/api/v1/ordering/orders", json={"name": "IMP"}, headers=headers)
    if r.status_code == 201:
        assert soap.global_sku not in {ln["global_sku"] for ln in r.json()["lines"]}

    r = client.post(
        f"/api/v1/ordering/vendors/{vendor_id}/orders",
        json={"quantities": {soap.global_sku: 1080}, "name": "Botanie July"},
        headers=headers,
    )
    assert r.status_code == 201, r.text
    dom_id = r.json()["order"]["id"]
    assert r.json()["order"]["order_type"] == "domestic"
    r = client.post(f"/api/v1/ordering/orders/{dom_id}/place", headers=headers)
    assert r.status_code == 200
    timeline = client.get(f"/api/v1/ordering/orders/{dom_id}/timeline", headers=headers).json()
    outbound = [m for m in timeline["emails"] if m["direction"] == "out"]
    assert outbound[0]["recipients"].startswith("orders@botanie.test")
    assert "1080" in outbound[0]["body"]  # per-vendor email lists item + qty

    # unknown SKU or someone else's product is refused
    r = client.post(
        f"/api/v1/ordering/vendors/{vendor_id}/orders",
        json={"quantities": {"NOPE": 5}},
        headers=headers,
    )
    assert r.status_code == 400


def test_canada_destination_seam(db, client):
    headers = _admin(db, client)
    _seed_import_products(db)
    r = client.post(
        "/api/v1/ordering/orders",
        json={"name": "CAN Q3", "destination": "CAN"},
        headers=headers,
    )
    assert r.status_code == 201
    order_id = r.json()["order"]["id"]
    assert r.json()["order"]["destination"] == "CAN"
    timeline = client.get(f"/api/v1/ordering/orders/{order_id}/timeline", headers=headers).json()
    assert any("USA→CAN" in e["note"] for e in timeline["events"])
    # destination rides the export
    r = client.get(f"/api/v1/ordering/orders/{order_id}/export.csv", headers=headers)
    assert ",CAN" in r.text.splitlines()[1]
    # and bad destinations are refused
    r = client.post(
        "/api/v1/ordering/orders", json={"name": "X", "destination": "MX"}, headers=headers
    )
    assert r.status_code == 400


def test_analogy_lifecycle(db, client):
    headers = _admin(db, client)
    lamp, _ = _seed_import_products(db)
    fresh = mk_product(db, "CA0023000009", "Copper Guru Puja Lamp Small", category="CONX", price=25)

    # suggest: no LLM key -> deterministic heuristic (shared name tokens)
    r = client.post(
        "/api/v1/ordering/analogies/suggest", json={"product_id": fresh.id}, headers=headers
    )
    assert r.status_code == 200
    suggestion = r.json()
    assert suggestion["analog_sku"] == lamp.global_sku
    assert suggestion["source"] == "heuristic"

    # confirm it -> analogy row; the review now forecasts by analogy
    r = client.post(
        "/api/v1/ordering/analogies",
        json={
            "product_id": fresh.id,
            "analog_product_id": suggestion["analog_product_id"],
            "rationale": suggestion["rationale"],
            "source": "llm",
        },
        headers=headers,
    )
    assert r.status_code == 201
    r = client.post("/api/v1/ordering/orders", json={"name": "A1"}, headers=headers)
    lines = {ln["global_sku"]: ln for ln in r.json()["lines"]}
    fresh_line = lines[fresh.global_sku]
    assert fresh_line["suggestion"]["forecast_method"] == "analogy"
    assert "analogy" in fresh_line["suggestion"]["flags"]
    assert fresh_line["suggested_sea_qty"] == 800  # borrowed from the lamp's 100/mo

    # once real history accumulates the analogy auto-graduates
    _seed_sales(db, fresh, 30, months=7)
    r = client.post("/api/v1/ordering/orders", json={"name": "A2"}, headers=headers)
    lines = {ln["global_sku"]: ln for ln in r.json()["lines"]}
    assert lines[fresh.global_sku]["suggestion"]["forecast_method"] != "analogy"
    r = client.get("/api/v1/ordering/analogies", headers=headers)
    assert r.json()[0]["status"] == "graduated"


def test_rules_overrides_change_the_engine(db, client):
    headers = _admin(db, client)
    _seed_import_products(db)
    r = client.put(
        "/api/v1/ordering/rules",
        json={"category_target_moh": {"CONX": 5}},
        headers=headers,
    )
    assert r.status_code == 200
    assert r.json()["effective"]["category_target_moh"]["CONX"] == 5.0
    r = client.post("/api/v1/ordering/orders", json={"name": "R2"}, headers=headers)
    lines = {ln["global_sku"]: ln for ln in r.json()["lines"]}
    assert lines[LAMP_SKU]["suggested_sea_qty"] == 500  # 5 months * 100/mo
    assert lines[DHOOP_SKU]["suggested_sea_qty"] == 400  # INCENSE untouched


def test_ordering_is_admin_only(db, client):
    mk_user(db, "wh@test.io", (Role.WAREHOUSE, None, None))
    headers = login(client, "wh@test.io")
    assert client.get("/api/v1/ordering/orders", headers=headers).status_code == 403
    assert (
        client.post("/api/v1/ordering/orders", json={"name": "X"}, headers=headers).status_code
        == 403
    )


def test_clothing_and_excluded_products_never_reach_the_review(db, client):
    headers = _admin(db, client)
    _seed_import_products(db)
    shirt = mk_product(db, "CL0000000001", "Adiyogi T-Shirt", category="Isha Life USA / Clothing")
    excluded = mk_product(db, "HO0009999999", "Retired Lamp", category="CONX")
    excluded.ordering_exclude = True
    db.commit()
    _seed_sales(db, shirt, 40)
    _seed_sales(db, excluded, 40)
    r = client.post("/api/v1/ordering/orders", json={"name": "F1"}, headers=headers)
    skus = {ln["global_sku"] for ln in r.json()["lines"]}
    assert shirt.global_sku not in skus
    assert excluded.global_sku not in skus


def test_odoo_sourcing_tags_steer_india_candidacy(db, client):
    """Products tagged in Odoo (synced into `Product.sourcing`): "Domestic"
    is a hard exclude from the India table — even with an India-shaped
    reference, even on the uploaded product list — and "India" makes a
    candidate of a product whose reference doesn't look India-shaped."""
    from app.ordering.inputs import import_candidates

    _admin(db, client)
    _seed_import_products(db)
    tagged_domestic = mk_product(db, "GL0000000404", "Glass Bottle (US-made)", category="Home")
    tagged_domestic.sourcing = "domestic"
    tagged_india = mk_product(db, "US-HW0100", "Herbal Handwash", category="Body Care")
    tagged_india.sourcing = "india"
    untagged_us = mk_product(db, "US-HW0200", "Hand Towel", category="Home")
    db.commit()

    ids = {p.id for p in import_candidates(db)}
    assert tagged_domestic.id not in ids  # India-shaped ref, Odoo says domestic
    assert tagged_india.id in ids  # US-shaped ref, Odoo says india
    assert untagged_us.id not in ids  # US-shaped ref, untagged: unchanged

    # the uploaded product list can't re-admit a Domestic-tagged product
    ids = {p.id for p in import_candidates(db, restrict_skus={"gl0000000404", "us-hw0100"})}
    assert tagged_domestic.id not in ids
    assert tagged_india.id in ids


def test_workbook_upload_generates_order(db, client):
    """The CSV fallback path: long-format sales history upload."""
    headers = _admin(db, client)
    lamp, _ = _seed_import_products(db)
    today = utcnow().date()
    rows = ["global_sku,year,month,units"]
    total = today.year * 12 + today.month - 1
    for back in range(1, 13):
        ordinal = total - back
        rows.append(f"{LAMP_SKU},{ordinal // 12},{ordinal % 12 + 1},200")
    csv_bytes = "\n".join(rows).encode()
    r = client.post(
        "/api/v1/ordering/orders/upload",
        files={"file": ("sales.csv", csv_bytes, "text/csv")},
        data={"name": "From CSV"},
        headers=headers,
    )
    assert r.status_code == 201, r.text
    detail = r.json()
    assert detail["order"]["snapshot_source"] == "csv"
    lines = {ln["global_sku"]: ln for ln in detail["lines"]}
    # the upload's 200/mo beats the snapshot's 100/mo
    assert lines[LAMP_SKU]["suggested_sea_qty"] == 1600


def test_orders_pin_their_snapshot(db, client):
    """DECISIONS.md: orders pin the snapshot they were computed from — later
    catalog/sales changes must not rewrite an existing draft."""
    headers = _admin(db, client)
    lamp, _ = _seed_import_products(db)
    r = client.post("/api/v1/ordering/orders", json={"name": "PIN"}, headers=headers)
    order_id = r.json()["order"]["id"]
    lamp.cost = 999  # catalog changes after the draft
    db.commit()
    detail = client.get(f"/api/v1/ordering/orders/{order_id}", headers=headers).json()
    lamp_line = next(ln for ln in detail["lines"] if ln["global_sku"] == LAMP_SKU)
    assert lamp_line["suggestion"]["unit_cost"] == 20.0  # frozen at creation
    assert detail["rules"]["sea_lead_months"] == 6  # rules frozen too
