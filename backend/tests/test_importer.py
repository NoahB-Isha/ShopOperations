from __future__ import annotations

from pathlib import Path

import pytest
from app.centers.importer import run_import
from app.models import Center, RoleAssignment, User, Zone
from sqlalchemy import select

US_HEADERS = [
    "City", "State", "Region", "City Coordinator Role", "Name", "Email", "Phone",
    "Contact added in WhattsApp group", "Address and email confirmed", "Address",
    "Stripe Terminal Name", "Stripe Terminal SN", "Zone", "Scanner",
    "IL Products for Sale or Display?", "Is Active as of Jan 2026?", "Notes",
    "Zone Coordinator", "Stripe Serial number", "Employee Pin codes",
    "Stripe Terminal tracking", "Active?",
]
CA_HEADERS = US_HEADERS


@pytest.fixture()
def workbook(tmp_path) -> Path:
    """A miniature of the real sheet, reproducing its messiness patterns."""
    import openpyxl

    wb = openpyxl.Workbook()

    us = wb.active
    us.title = "US"
    us.append(US_HEADERS)

    def us_row(city, state, name="", email="", phone="", zone="", jan="", notes="",
               coord="", active="", terminal=""):
        us.append([city, state, "Region", "Shoppe", name, email, phone, "Yes", "Yes",
                   "123 St", terminal, "", zone, "", "Sale", jan, notes, coord, "", "", "", active])

    us_row("Austin", "Texas", "Sachi M", "smutluru@yahoo.com", "(512) 937-8219",
           zone="1.0", coord="Lili", active="Yes",
           notes="Austin shares Products with San Antonio")
    us_row("San Antonio", "Texas", "Akshay V", "akshay@example.com", "(210) 555-0101",
           zone="1.0", coord="Lili", active="Yes",
           notes="San Antonio shares Products with Austin")
    # messy: missing email AND phone, ambiguous active ('?')
    us_row("Cincinnati", "Ohio", jan="?")
    # messy: phone only, explicitly inactive
    us_row("Rochester", "New York", "Priya R", "", "(585) 555-0177", active="No")
    # messy: stray space in email, temporary note, no zone at all
    us_row("Phoenix", "Arizona", "Lomush M", "tiba@bellsouth. net", "(561) 818-1829",
           active="Yes", notes="Temporary")

    ca = wb.create_sheet("canada new")
    ca.append(CA_HEADERS)
    ca.append(["Toronto-West", "Ontario", "Canada", "IE IL-WEST", "Sneha S",
               "sneha@example.ca", "416-684-2298", "Yes", "Yes", "5816 Street",
               "ST24-361", "WSC50S-0AA", "Toronto", "No", "Yes", "Yes", "", "", "", "", "", ""])
    # continuation row: extra terminal for the previous center, no city
    ca.append(["", "", "", "", "Sneha S", "", "", "", "", "",
               "ST13-212", "WSC50S-00K", "Toronto", "No", "", "", "", "", "", "", "", ""])
    # same center again = a second contact
    ca.append(["Toronto-West", "Ontario", "Canada", "IL Satsang West", "Minh L",
               "", "647-649-7746", "Yes", "", "", "", "", "", "", "Yes", "", "", "", "", "", "", ""])
    # a center with no reachable contact at all
    ca.append(["Toronto-London", "Ontario", "Canada", "", "", "", "", "", "", "",
               "", "", "", "", "Yes", "", "", "", "", "", "", ""])

    legacy = wb.create_sheet("Old")
    legacy.append(["City", "Region"])
    legacy.append(["Atlanta", "Midwest"])
    legacy2 = wb.create_sheet("Canada (old)")
    legacy2.append(["City", "Region"])
    legacy2.append(["Edmonton", "West Canada"])

    path = tmp_path / "coordinators.xlsx"
    wb.save(path)
    return path


def test_import_zones_centers_and_flags(db, workbook):
    report = run_import(db, workbook, apply=True)

    assert report.sheets_processed == ["US", "canada new"]
    assert set(report.sheets_skipped) == {"Old", "Canada (old)"}

    zones = {z.name for z in db.scalars(select(Zone))}
    assert zones == {"Zone 1 (Lili)", "Canada"}

    centers = {c.name: c for c in db.scalars(select(Center))}
    assert len(centers) == 7

    austin, sa = centers["Austin"], centers["San Antonio"]
    assert austin.is_active and austin.zone.name == "Zone 1 (Lili)"
    # shared product set: both sides carry the same group label
    assert austin.shared_product_group == sa.shared_product_group is not None

    cin = centers["Cincinnati"]
    assert not cin.is_active  # ambiguous -> inactive, flagged, never guessed
    assert "ambiguous_active" in cin.followup_reasons
    assert "no_reachable_contact" in cin.followup_reasons
    assert cin.needs_followup

    roch = centers["Rochester"]
    assert not roch.is_active
    assert "contact_missing_email" in roch.followup_reasons

    phx = centers["Phoenix"]
    assert "temporary" in phx.followup_reasons and "no_zone" in phx.followup_reasons
    assert phx.contacts[0].email == "tiba@bellsouth.net"  # stray space repaired

    tw = centers["Toronto-West"]
    assert tw.country == "CA" and tw.zone.name == "Canada"
    assert tw.is_active  # canada sheet defaults active
    assert len(tw.contacts) == 2  # merged second row
    assert "ST13-212" in tw.stripe_terminal_serial or "ST13-212" in tw.stripe_terminal_name

    tl = centers["Toronto-London"]
    assert "no_reachable_contact" in tl.followup_reasons


def test_import_creates_orderer_users_only_for_reachable_active(db, workbook):
    report = run_import(db, workbook, apply=True)
    users = db.scalars(select(User)).all()
    emails = {u.email for u in users if u.email}
    assert "smutluru@yahoo.com" in emails
    assert "sneha@example.ca" in emails
    phones = {u.phone for u in users if u.phone}
    assert "+16476497746" in phones  # phone-only contact still gets an account
    # inactive Rochester and contactless Cincinnati produce no users
    assert not any(u.email and "priya" in (u.display_name or "").lower() for u in users)
    assert report.users_created == len(users)

    sachi = next(u for u in users if u.email == "smutluru@yahoo.com")
    roles = db.scalars(select(RoleAssignment).where(RoleAssignment.user_id == sachi.id)).all()
    assert [r.role for r in roles] == ["center_orderer"]


def test_import_is_idempotent(db, workbook):
    first = run_import(db, workbook, apply=True)
    second = run_import(db, workbook, apply=True)
    assert second.centers_created == 0
    assert second.centers_updated == first.centers_created
    assert second.users_created == 0
    assert db.scalar(select(Center).where(Center.name == "Austin")) is not None
    assert len(db.scalars(select(Center)).all()) == 7
    tw = db.scalar(select(Center).where(Center.name == "Toronto-West"))
    assert len(tw.contacts) == 2  # contacts not duplicated


def test_preview_mode_writes_nothing(db, workbook):
    report = run_import(db, workbook, apply=False)
    assert report.centers_parsed == 7
    assert not report.applied
    assert db.scalars(select(Center)).all() == []


def test_csv_roster_imports_like_a_one_sheet_workbook(db, tmp_path):
    """An admin uploads whatever their machine has. A CSV is one sheet, so the
    Zone column carries the zone split the workbook did with tabs — which is
    what the US sheets already do."""
    csv_path = tmp_path / "roster.csv"
    csv_path.write_text(
        ",".join(US_HEADERS)
        + "\n"
        + ",".join(
            [
                "Testville", "Ohio", "Midwest", "Shoppe", "Pat Example",
                "pat@example.org", "555-0100", "", "", "", "WPC-TEST-1", "SN-1",
                # the Zone column carries the CODE, exactly as the workbook's
                # US sheets do — `_col` fuzzy-matches headers, so "Zone
                # Coordinator" resolves to "Zone" and only the code path runs
                "3", "", "", "", "csv probe", "Ravi", "", "", "", "Yes",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    report = run_import(db, csv_path, apply=True, create_users=True)
    assert report.sheets_processed == ["roster"]  # the file name becomes the sheet

    center = db.scalar(select(Center).where(Center.name == "Testville"))
    assert center is not None
    assert center.state == "Ohio" and center.is_active is True
    assert center.stripe_terminal_name == "WPC-TEST-1"
    db.refresh(center)  # zone_id was just written; load the relationship fresh
    assert center.zone is not None and center.zone.name == "Zone 3 (Ravi)"
    assert [c.email for c in center.contacts] == ["pat@example.org"]


def test_unsupported_roster_file_is_refused_by_name(db, tmp_path):
    """Say which formats work, rather than dying inside a parser."""
    bad = tmp_path / "roster.numbers"
    bad.write_text("nope", encoding="utf-8")
    with pytest.raises(ValueError, match="xlsx or .csv"):
        run_import(db, bad, apply=False)
