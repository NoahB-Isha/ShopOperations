"""THE parity test — the ordering engine driven by the spec-of-record
workbook's own inputs.

For every fully-numeric row of the SEA sheet in
`docs/reference/USA INV CHK.xlsx` (281 rows in the committed copy), feed the
engine the row's MON SALES, MOH, target MOH and the six incoming-MOH columns,
and assert our SEA QTY / AIR QTY match the workbook's computed values within
rounding. This is the acceptance criterion for the Phase-4 math: the workbook
is the provable baseline the smart forecast generalises from.

SEA sheet columns used: A name, F MON SALES, H MOH, J MTHS REQ (target),
Q SEA QTY, S AIR QTY, V..AA the +1..+6 MOH INC columns.
"""

from __future__ import annotations

import warnings
from pathlib import Path

import pytest
from app.ordering.engine import ProductInput, SkuSnapshot, suggest_one
from app.ordering.rules import OrderingRules

WORKBOOK = Path(__file__).resolve().parents[2] / "docs" / "reference" / "USA INV CHK.xlsx"

pytestmark = pytest.mark.skipif(
    not WORKBOOK.exists(), reason="spec workbook not present at docs/reference/"
)


def _load_sea_rows():
    import openpyxl

    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb["SEA"]
    rows = []
    for r in range(2, ws.max_row + 1):
        def g(i, _r=r):
            return ws.cell(_r, i).value

        mon, moh, target = g(6), g(8), g(10)
        seaqty, airqty = g(17), g(19)
        inc = [g(i) for i in range(22, 28)]
        vals = [mon, moh, target, seaqty, airqty] + inc
        if all(isinstance(v, int | float) for v in vals) and mon and mon > 0:
            rows.append((g(1), mon, moh, target, seaqty, airqty, inc))
    return rows


def test_engine_matches_every_numeric_sea_row():
    rules = OrderingRules()
    rows = _load_sea_rows()
    assert len(rows) >= 250, f"expected the full SEA sheet, got {len(rows)} rows"
    mismatches = []
    for name, mon, moh, target, seaqty, airqty, inc in rows:
        snap = SkuSnapshot(
            product=ProductInput(global_sku="X", category="X", cost=1, retail_price=2,
                                 target_moh_override=target),
            on_hand=moh * mon,
            avg_monthly_sales=mon,
            incoming_units_by_month=[m * mon for m in inc],
            forecast=None,
        )
        s = suggest_one(snap, rules)
        if abs(s.suggested_sea_qty - seaqty) > 1.0 or abs(s.suggested_air_qty - airqty) > 1.0:
            mismatches.append((name, s.suggested_sea_qty, seaqty, s.suggested_air_qty, airqty))
    assert not mismatches, f"{len(mismatches)}/{len(rows)} rows differ: {mismatches[:5]}"
