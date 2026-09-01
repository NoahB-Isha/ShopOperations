"""Security hardening regressions — one test per closed hole.

These cover inputs that reach a file the India team opens, a response header,
or a quantity on a real Odoo draft: spreadsheet-formula injection in exports,
CRLF/content-type injection in downloads, non-finite or unbounded quantities,
and unbounded request bodies.
"""
from __future__ import annotations

import io
from pathlib import Path

import openpyxl
import pytest
from app.counting.router import CountLineIn
from app.downloads import attachment_headers, safe_filename
from app.models import (
    OdooWriteOutcome,
    OrderAttachment,
    PalletTransfer,
    Product,
    PurchaseOrder,
    Role,
)
from app.odoo.simulator import OdooSimulator
from app.odoo.writer import OdooWriter, WriterValidationError
from app.ordering.export import EXPORT_COLUMNS, _safe_cell, rows_to_csv, rows_to_xlsx
from app.ordering.inputs import build_bundle_from_workbook
from app.ordering.rules import OrderingRules
from app.sync.runner import run_domain
from pydantic import ValidationError
from sqlalchemy import func, select

from .util import login, mk_user


# ------------------------------------------------------ C-10 download headers
def test_safe_filename_strips_interior_crlf() -> None:
    """The mailbox vector: an RFC 2231-decoded attachment name carrying a
    header break, reachable with no app account."""
    name = safe_filename("invoice\r\nX-Injected: yes.pdf")
    assert "\r" not in name and "\n" not in name
    assert name == "invoiceX-Injected: yes.pdf"

    header = attachment_headers("invoice\r\nX-Injected: yes.pdf")["Content-Disposition"]
    assert "\r" not in header and "\n" not in header
    assert header.startswith('attachment; filename="')
    # a name with no usable characters still yields a header, never an empty one
    assert 'filename="download"' in attachment_headers("\r\n\t")["Content-Disposition"]


def test_stored_html_attachment_downloads_as_octet_stream(db, client) -> None:
    """A text/html attachment out of an email must download as opaque bytes —
    never render inside the app's origin — and its filename must not be able
    to inject a header."""
    mk_user(db, "admin@sec.io", (Role.ADMIN, None, None))
    headers = login(client, "admin@sec.io")
    order = PurchaseOrder(name="SEC-1", reference="ILAPP-PO-SEC1")
    db.add(order)
    db.flush()
    attachment = OrderAttachment(
        order_id=order.id,
        source="email",
        filename="reply\r\nX-Injected: yes.html",
        content_type="text/html",
        size_bytes=9,
        data=b"<script>alert(1)</script>",
    )
    db.add(attachment)
    db.commit()

    r = client.get(
        f"/api/v1/ordering/orders/{order.id}/attachments/{attachment.id}/download",
        headers=headers,
    )
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == "application/octet-stream"
    assert "x-injected" not in {k.lower() for k in r.headers}
    disposition = r.headers["content-disposition"]
    assert "\r" not in disposition and "\n" not in disposition
    assert "reply" in disposition


# -------------------------------------------------- C-3 formula injection
def test_safe_cell_escapes_formulas_and_leaves_numbers() -> None:
    assert _safe_cell("=cmd|'/c calc'!A1") == "'=cmd|'/c calc'!A1"
    assert _safe_cell("@SUM(A1:A2)") == "'@SUM(A1:A2)"
    assert _safe_cell("-1+1") == "'-1+1"
    assert _safe_cell("Copper Bottle") == "Copper Bottle"  # ordinary text untouched
    assert _safe_cell(42) == 42 and isinstance(_safe_cell(42), int)
    assert _safe_cell(3.5) == 3.5 and _safe_cell(None) is None


def _one_export_row(name: str) -> dict:
    row = {key: "" for key, _ in EXPORT_COLUMNS}
    row.update({"name": name, "final_sea_qty": 12, "final_air_qty": 0})
    return row


def test_exports_neutralize_a_formula_name() -> None:
    csv_text = rows_to_csv([_one_export_row("=HYPERLINK(\"http://evil\",\"click\")")])
    assert "'=HYPERLINK" in csv_text

    wb = openpyxl.load_workbook(io.BytesIO(rows_to_xlsx([_one_export_row("=1+1")])))
    ws = wb["Order"]
    name_col = 1 + [k for k, _ in EXPORT_COLUMNS].index("name")
    cell = ws.cell(2, name_col)
    assert str(cell.value).startswith("'")
    assert cell.quotePrefix is True  # Excel's own "this is text" marker
    qty_col = 1 + [k for k, _ in EXPORT_COLUMNS].index("final_sea_qty")
    assert ws.cell(2, qty_col).value == 12  # numeric columns keep native types
    wb.close()


# ------------------------------------------- C-11 bounded, finite quantities
def test_counting_rejects_non_finite_and_negative_counts() -> None:
    """The one door counted numbers enter through (the counting page) keeps
    the NaN/inf/bounds guard — NaN slipped past `<= 0` checks before."""
    with pytest.raises(ValidationError):
        CountLineIn(product_id=1, counted_qty=float("nan"))
    with pytest.raises(ValidationError):
        CountLineIn(product_id=1, counted_qty="NaN")  # type: ignore[arg-type]
    with pytest.raises(ValidationError):
        CountLineIn(product_id=1, counted_qty=float("inf"))
    with pytest.raises(ValidationError):
        CountLineIn(product_id=1, counted_qty=-1)
    with pytest.raises(ValidationError):
        CountLineIn(product_id=1, counted_qty=1_000_001)
    assert CountLineIn(product_id=1, counted_qty=7).counted_qty == 7


def test_writer_rejects_non_finite_quantities(db, settings_env) -> None:
    """NaN fails every comparison, so the old `qty <= 0` guard passed it
    straight onto the move payload."""
    sim = OdooSimulator(settings_env.fixtures_path, read_only=False)
    run_domain(db, settings_env, "products", conn=sim, trigger="manual")
    run_domain(db, settings_env, "stock", conn=sim, trigger="manual")
    product = db.scalar(select(Product).where(Product.global_sku == "CA0023000009"))
    writer = OdooWriter(db, settings_env, conn=sim)

    for bad in (float("nan"), float("inf"), float("-inf"), 0.0, -5.0):
        with pytest.raises(WriterValidationError):
            writer.create_internal_transfer(
                source_key="bwhse",
                dest_key="floor",
                lines=[{"product_id": product.id, "qty": bad}],
            )
        with pytest.raises(WriterValidationError):
            writer.create_inventory_reduction(product_id=product.id, qty=bad)


# ------------------------------------------------ C-9d bounded workbook read
WORKBOOK = Path(__file__).resolve().parents[2] / "docs" / "reference" / "USA INV CHK.xlsx"


@pytest.mark.skipif(not WORKBOOK.exists(), reason="spec workbook not present")
def test_workbook_upload_still_reads_every_column_in_read_only_mode(db) -> None:
    """read_only=True bounds the memory an uploaded workbook can claim, but it
    makes the sheet forward-only — the row loop streams tuples instead of
    calling ws.cell(). This proves the same columns still land, against the
    real spec workbook."""
    bundle = build_bundle_from_workbook(db, OrderingRules(), WORKBOOK.read_bytes())
    assert len(bundle.snapshots) >= 250  # 289 numeric SEA rows in the committed copy
    first = bundle.snapshots[0]
    assert first.product.global_sku and first.product.name  # cols 2 and 1
    assert first.avg_monthly_sales > 0  # col 6
    assert len(first.incoming_units_by_month) == OrderingRules().horizon  # cols 22..27
    assert all(v >= 0 for v in first.incoming_units_by_month)

    empty = build_bundle_from_workbook(db, OrderingRules(), _workbook_without_sea())
    assert empty.snapshots == [] and "no SEA sheet" in empty.warnings[0]


def _workbook_without_sea() -> bytes:
    wb = openpyxl.Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ------------------------------------------- C-13 one open pallet at a time
def test_second_pallet_is_refused_while_one_awaits_validation(db, settings_env) -> None:
    """Each pallet carries its own ILAPP-PLT- reference, so the writer's
    origin-keyed dedupe cannot catch a double 'Send all' — the guard lives in
    create_pallet, and it must fire before anything is read or written."""
    from app.transfers.pallet import create_pallet, open_pallet

    assert open_pallet(db) is None
    db.add(
        PalletTransfer(
            picking_reference="ILAPP-PLT-AAAA",
            picking_status=OdooWriteOutcome.CREATED.value,
            odoo_picking_id=9001,
            odoo_picking_name="III/INT/09001",
            lines=[{"product_id": 1, "sku": "X", "name": "X", "qty": 1}],
        )
    )
    db.commit()
    assert open_pallet(db) is not None

    with pytest.raises(ValueError, match="III/INT/09001"):
        create_pallet(db, settings_env, actor_user_id=None)
    assert db.scalar(select(func.count()).select_from(PalletTransfer)) == 1


# ------------------------------------------------- C-12 bounded list bodies
def test_transfer_request_with_too_many_lines_is_rejected(db, client) -> None:
    mk_user(db, "floor@sec.io", (Role.SHOPPE_FLOOR, None, None))
    headers = login(client, "floor@sec.io")
    lines = [{"product_id": i, "qty": 1} for i in range(1, 502)]  # 501 lines
    r = client.post(
        "/api/v1/transfer-requests", json={"lines": lines}, headers=headers
    )
    assert r.status_code == 422, r.text
    r = client.put(
        "/api/v1/transfer-requests/1/lines", json={"lines": lines}, headers=headers
    )
    assert r.status_code == 422, r.text
