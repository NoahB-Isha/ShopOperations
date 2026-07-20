"""Order export — CSV and XLSX with the agreed ORDER LIST columns, handed
straight to the India team / customs broker. Quantities are the FINAL
(override- and event-aware) values; everything else comes from the frozen
per-line suggestion so a later catalog change never rewrites a placed order.
"""

from __future__ import annotations

import csv
import io
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models import PurchaseOrder, PurchaseOrderLine

EXPORT_COLUMNS = [
    ("us_sku", "US SKU"),
    ("name", "NAME"),
    ("global_sku", "GLOBAL SKU"),
    ("category", "CATEGORY"),
    ("final_sea_qty", "SEA QTY"),
    ("final_air_qty", "AIR QTY"),
    ("unit_weight", "UNIT WEIGHT (G)"),
    ("unit_cost", "UNIT COST (COGS)"),
    ("retail_price", "RETAIL"),
    ("margin", "MARGIN"),
    ("hsn_code", "HSN"),
    ("air_shipping_cost", "AIR SHIPPING COST"),
    ("profit_lost_by_air", "PROFIT LOST BY AIR"),
    ("target_moh", "TARGET MOH"),
    ("case_size", "CASE"),
    ("flags", "FLAGS"),
    ("destination", "DESTINATION"),
]


def export_rows(order: PurchaseOrder) -> list[dict[str, Any]]:
    """Only lines actually being ordered (a final quantity > 0, not
    discontinued), sorted the way the workbook reads."""
    rows = []
    for line in order.lines:
        if line.line_status == "discontinued":
            continue
        if (line.final_sea_qty or 0) <= 0 and (line.final_air_qty or 0) <= 0:
            continue
        s = line.suggestion_json or {}
        unit_cost = float(s.get("unit_cost") or 0)
        margin = float(s.get("margin") or 0)
        rows.append(
            {
                "us_sku": s.get("us_sku", ""),
                "name": s.get("name", ""),
                "global_sku": line.global_sku,
                "category": s.get("category", ""),
                "final_sea_qty": line.final_sea_qty,
                "final_air_qty": line.final_air_qty,
                "unit_weight": s.get("unit_weight_g"),
                "unit_cost": unit_cost,
                "retail_price": s.get("retail_price", 0),
                "margin": margin,
                "hsn_code": s.get("hsn_code", ""),
                "air_shipping_cost": round(unit_cost * line.final_air_qty, 2),
                "profit_lost_by_air": round(margin * line.final_air_qty, 2),
                "target_moh": line.target_moh_used,
                "case_size": line.case_size,
                "flags": ", ".join(s.get("flags") or []),
                "destination": order.destination,
            }
        )
    rows.sort(key=lambda r: (str(r["category"]), str(r["us_sku"]), str(r["global_sku"])))
    return rows


def rows_to_csv(rows: list[dict[str, Any]]) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([header for _, header in EXPORT_COLUMNS])
    for r in rows:
        writer.writerow(["" if r.get(k) is None else r.get(k) for k, _ in EXPORT_COLUMNS])
    return buf.getvalue()


def rows_to_xlsx(rows: list[dict[str, Any]], order_name: str = "Order") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (order_name or "Order")[:31]
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF")
    for c, (_, header) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(1, c, header)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    flag_fill = PatternFill("solid", fgColor="FFF2CC")
    for ri, r in enumerate(rows, start=2):
        for c, (k, _) in enumerate(EXPORT_COLUMNS, start=1):
            ws.cell(ri, c, "" if r.get(k) is None else r.get(k))
        if r.get("flags"):
            for c in range(1, len(EXPORT_COLUMNS) + 1):
                ws.cell(ri, c).fill = flag_fill
    widths = [12, 42, 16, 18, 9, 9, 14, 15, 9, 10, 12, 16, 17, 11, 7, 24, 12]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def vendor_email_lines(order: PurchaseOrder) -> list[tuple[str, int]]:
    """(item name, qty) pairs for a domestic vendor email body."""
    out = []
    for line in order.lines:
        qty = (line.final_sea_qty or 0) + (line.final_air_qty or 0)
        if qty <= 0 or line.line_status == "discontinued":
            continue
        s = line.suggestion_json or {}
        out.append((str(s.get("name") or line.global_sku), qty))
    return out


def line_qty(line: PurchaseOrderLine) -> int:
    return (line.final_sea_qty or 0) + (line.final_air_qty or 0)
