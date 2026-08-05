"""Engine inputs — where the app's own Odoo snapshot (or an uploaded
workbook/CSV) becomes `SkuSnapshot`s for the suggestion engine.

Snapshot path (the normal one):
  * candidates: active, Odoo-sourced, stock-tracked products that are not
    clothing (project brief: out of scope), not `ordering_exclude`d, and are
    India imports — internal reference matching ^[A-Za-z]{2}\\d{10}$ (verified
    instance fact), an explicitly assigned India vendor, or an "India" product
    tag in Odoo (`Product.sourcing`). A product assigned a US/CA vendor OR
    tagged "Domestic" in Odoo is domestic and stays off the India table.
  * on-hand: BWHSE + FLOOR + STAGING stock levels summed (whole-campus stock,
    matching the workbook's INV OH; city-center consignment stock excluded).
  * sales: `sales_monthly` (24-month history), channels summed, SPARSE — only
    months that actually sold units appear, so avg = units per selling month,
    the workbook's sell-through-velocity semantics. The current (incomplete)
    month is excluded so a mid-month refresh doesn't deflate velocity.
  * incoming: `incoming_moves` bucketed into projection months 1..H by
    expected_date (missing dates land in month 1 — such moves are typically
    already assigned; treating them as imminent avoids re-ordering them).

New products: an ACTIVE ForecastAnalogy supplies demand when real history is
thinner than the low-confidence bar — either the analog product's own
forecast or a flat manual estimate, always labelled method="analogy" so the
review screen shows forecast-by-analogy honestly. Once real history reaches
`analogy_graduation_months` the analogy is ignored (and reported for the
service layer to mark graduated).

Upload fallback: the USA INV CHK workbook's SEA sheet (full standalone
inputs, per-row target respected) or a long-format sales CSV
(`global_sku,year,month,units[,is_stockout]` — stock/incoming still from the
app snapshot). Zero Odoo required for either.
"""

from __future__ import annotations

import csv
import io
import re
import warnings as _warnings
from collections.abc import Sequence
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any

from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from ..models import (
    SOURCING_DOMESTIC,
    SOURCING_INDIA,
    AnalogyStatus,
    ForecastAnalogy,
    IncomingMove,
    Product,
    SalesMonthly,
    StockLevel,
    Vendor,
    VendorKind,
    not_blacklisted,
    not_clothing,
    utcnow,
)
from .engine import ProductInput, SkuSnapshot
from .forecasting import (
    METHOD_ANALOGY,
    Forecast,
    MonthlySalesSeries,
    MonthPoint,
    flat_forecast,
    forecast_demand,
)
from .rules import AIR_ONLY_TAGS, BULK_CYCLE_TAGS, EXPIRY_TAGS, OrderingRules

INDIA_REF_RE = re.compile(r"^[A-Za-z]{2}\d{10}$")

# staging2 = warehouse consolidation stock, still owned — counts for purchasing
ON_HAND_LOCATION_KEYS = ("bwhse", "floor", "staging", "staging2")


def is_india_ref(code: str | None) -> bool:
    return bool(code and INDIA_REF_RE.match(code.strip()))


@dataclass
class SnapshotBundle:
    """Everything one review/order generation was computed from."""

    snapshots: list[SkuSnapshot] = field(default_factory=list)
    snapshot_at: datetime = field(default_factory=utcnow)
    source: str = "odoo"  # odoo | workbook | csv
    warnings: list[str] = field(default_factory=list)
    # analogy ids whose products now have enough real history to graduate
    graduable_analogy_ids: list[int] = field(default_factory=list)


# ------------------------------------------------------------ product -> input
def product_to_input(product: Product, vendor: Vendor | None = None) -> ProductInput:
    tags = {t.tag for t in product.tags}
    vendor = vendor if vendor is not None else product_vendor(product)
    is_domestic = bool(vendor and vendor.kind != VendorKind.INDIA.value)
    return ProductInput(
        global_sku=product.global_sku,
        name=product.name,
        us_sku=product.us_sku,
        category=product.category,
        # case_size 1 is the column default and means "unset" in practice —
        # category defaults (BLOOM 32) apply until an admin sets a real value.
        case_size=product.case_size if product.case_size and product.case_size > 1 else None,
        unit_weight_g=product.unit_weight_g,
        hsn_code=product.hsn_code,
        cost=float(product.cost or 0),
        retail_price=float(product.retail_price or 0),
        target_moh_override=product.target_moh_override,
        is_domestic=is_domestic,
        moq=product.moq,
        air_only=any(t in tags for t in AIR_ONLY_TAGS),
        sea_only="sea_only" in tags,
        bulk_cycle=any(t in tags for t in BULK_CYCLE_TAGS),
        expiry_sensitive=any(t in tags for t in EXPIRY_TAGS),
    )


def product_vendor(product: Product) -> Vendor | None:
    if product.vendor_id is None:
        return None
    return Session.object_session(product).get(Vendor, product.vendor_id)  # type: ignore[union-attr]


def import_candidates(db: Session, restrict_skus: set[str] | None = None) -> list[Product]:
    """Products eligible for the India review table.

    `restrict_skus` (the admin's uploaded product list) is authoritative when
    present: listed products are included even when their reference doesn't
    look India-shaped — the buyer curated the list. Two exclusions outrank
    even the list, because both are explicit human declarations that the
    product is bought domestically: a domestic-vendor assignment (the vendor
    flow owns it) and a "Domestic" product tag in Odoo (`Product.sourcing`,
    synced). Symmetrically, an "India" tag makes a product a candidate even
    when its reference isn't India-shaped."""
    india_vendor_ids = [
        vid for vid, in db.execute(select(Vendor.id).where(Vendor.kind == VendorKind.INDIA.value))
    ]
    products = (
        db.execute(
            select(Product)
            .options(selectinload(Product.tags))
            .where(
                Product.is_active.is_(True),
                Product.source == "odoo",
                Product.is_stock_tracked.is_(True),
                Product.ordering_exclude.is_(False),
                not_clothing(),
                not_blacklisted(),
            )
            .order_by(Product.category, Product.name)
        )
        .scalars()
        .all()
    )
    restrict = {s.strip().lower() for s in restrict_skus} if restrict_skus is not None else None
    out = []
    for p in products:
        if p.vendor_id is not None and p.vendor_id not in india_vendor_ids:
            continue  # assigned to a domestic vendor -> vendor flow
        if p.sourcing == SOURCING_DOMESTIC:
            continue  # tagged Domestic in Odoo -> never on the India table
        if restrict is not None:
            if p.global_sku.lower() in restrict or (
                p.odoo_internal_ref and p.odoo_internal_ref.lower() in restrict
            ):
                out.append(p)
            continue
        if p.vendor_id is not None or p.sourcing == SOURCING_INDIA:
            out.append(p)  # india vendor or tagged India in Odoo
            continue
        if is_india_ref(p.global_sku) or is_india_ref(p.odoo_internal_ref):
            out.append(p)
    return out


# ------------------------------------------------------------- app snapshot
def _month_index_from_today(expected: date | None, today: date) -> int:
    """Projection month (1-based) an incoming move lands in."""
    if expected is None:
        return 1  # assigned moves with no date are typically imminent
    delta = (expected.year * 12 + expected.month) - (today.year * 12 + today.month)
    return max(1, delta + 1)


def _series_from_monthly_rows(
    rows: list[tuple[int, int, float]], today: date
) -> MonthlySalesSeries:
    """Sparse per-month totals -> series. The current month is dropped
    (incomplete); months are merged across channels upstream."""
    pts = [
        MonthPoint(year=y, month=m, units=u)
        for (y, m, u) in sorted(rows)
        if u > 0 and not (y == today.year and m == today.month)
    ]
    return MonthlySalesSeries(points=pts)


def _load_stock(db: Session, ids: list[int]) -> dict[int, float]:
    stock: dict[int, float] = {}
    for pid, qty in db.execute(
        select(StockLevel.product_id, StockLevel.qty).where(
            StockLevel.product_id.in_(ids),
            StockLevel.location_key.in_(ON_HAND_LOCATION_KEYS),
        )
    ):
        stock[pid] = stock.get(pid, 0.0) + float(qty or 0)
    return stock


def _load_monthly(db: Session, ids: list[int]) -> dict[int, dict[tuple[int, int], float]]:
    monthly: dict[int, dict[tuple[int, int], float]] = {}
    for pid, y, m, units in db.execute(
        select(
            SalesMonthly.product_id, SalesMonthly.year, SalesMonthly.month, SalesMonthly.units
        ).where(SalesMonthly.product_id.in_(ids))
    ):
        buckets = monthly.setdefault(pid, {})
        buckets[(y, m)] = buckets.get((y, m), 0.0) + float(units or 0)
    return monthly


def _load_incoming(
    db: Session, ids: list[int], rules: OrderingRules, today: date
) -> dict[int, list[float]]:
    incoming: dict[int, list[float]] = {}
    for pid, qty, expected in db.execute(
        select(IncomingMove.product_id, IncomingMove.qty, IncomingMove.expected_date).where(
            IncomingMove.product_id.in_(ids)
        )
    ):
        idx = _month_index_from_today(expected, today)
        if idx > rules.horizon:
            continue  # lands beyond the projection window
        arr = incoming.setdefault(pid, [0.0] * rules.horizon)
        arr[idx - 1] += float(qty or 0)
    return incoming


def snapshots_for_products(
    db: Session,
    rules: OrderingRules,
    products: Sequence[Product],
    vendors: dict[int, Vendor],
    today: date | None = None,
) -> list[SkuSnapshot]:
    """SkuSnapshots for an explicit product list (the domestic-vendor path —
    no analogies, no candidate filtering)."""
    today = today or utcnow().date()
    ids = [p.id for p in products]
    stock = _load_stock(db, ids)
    monthly = _load_monthly(db, ids)
    incoming = _load_incoming(db, ids, rules, today)
    snaps: list[SkuSnapshot] = []
    for p in products:
        series = _series_from_monthly_rows(
            [(y, m, u) for (y, m), u in monthly.get(p.id, {}).items()], today
        )
        forecast = (
            forecast_demand(series, rules.horizon, rules.forecast) if series.points else None
        )
        vendor = vendors.get(p.vendor_id) if p.vendor_id else None
        snaps.append(
            SkuSnapshot(
                product=product_to_input(p, vendor),
                on_hand=stock.get(p.id, 0.0),
                avg_monthly_sales=forecast.baseline if forecast else 0.0,
                incoming_units_by_month=incoming.get(p.id, [0.0] * rules.horizon),
                forecast=forecast,
                units_sold=sum(pt.units for pt in series.points),
                months_active=len(series.useable()),
            )
        )
    return snaps


def build_snapshot_bundle(
    db: Session,
    rules: OrderingRules,
    today: date | None = None,
    restrict_skus: set[str] | None = None,
) -> SnapshotBundle:
    """SkuSnapshots for every import candidate, from the app's own snapshot."""
    today = today or utcnow().date()
    bundle = SnapshotBundle(source="odoo")
    products = import_candidates(db, restrict_skus)
    if not products:
        bundle.warnings.append(
            "no products from the uploaded product list are import-eligible"
            if restrict_skus is not None
            else "no import-eligible products in the catalog"
        )
        return bundle
    ids = [p.id for p in products]

    stock = _load_stock(db, ids)
    monthly = _load_monthly(db, ids)
    incoming = _load_incoming(db, ids, rules, today)

    analogies = {
        a.product_id: a
        for a in db.execute(
            select(ForecastAnalogy).where(ForecastAnalogy.status == AnalogyStatus.ACTIVE.value)
        ).scalars()
    }
    products_by_id = {p.id: p for p in products}
    vendors = {v.id: v for v in db.execute(select(Vendor)).scalars()}

    for p in products:
        series = _series_from_monthly_rows(
            [(y, m, u) for (y, m), u in monthly.get(p.id, {}).items()], today
        )
        n_real = len(series.useable())
        units_sold = sum(pt.units for pt in series.points)
        forecast: Forecast | None = None
        analogy = analogies.get(p.id)

        if analogy and n_real >= rules.forecast.analogy_graduation_months:
            bundle.graduable_analogy_ids.append(analogy.id)
            analogy = None  # enough real history — the analogy retires

        if analogy and n_real < rules.forecast.low_confidence_months:
            forecast = _analogy_forecast(
                analogy, products_by_id, monthly, rules, today, bundle.warnings, db
            )
        if forecast is None and series.points:
            forecast = forecast_demand(series, rules.horizon, rules.forecast)

        avg = forecast.baseline if forecast else 0.0
        vendor = vendors.get(p.vendor_id) if p.vendor_id else None
        bundle.snapshots.append(
            SkuSnapshot(
                product=product_to_input(p, vendor),
                on_hand=stock.get(p.id, 0.0),
                avg_monthly_sales=avg,
                incoming_units_by_month=incoming.get(p.id, [0.0] * rules.horizon),
                forecast=forecast,
                units_sold=units_sold,
                months_active=n_real,
            )
        )
    return bundle


def _analogy_forecast(
    analogy: ForecastAnalogy,
    products_by_id: dict[int, Product],
    monthly: dict[int, dict[tuple[int, int], float]],
    rules: OrderingRules,
    today: date,
    warn: list[str],
    db: Session,
) -> Forecast | None:
    """Demand for a new product borrowed from its analog (or a flat manual
    estimate). Always labelled method='analogy'."""
    if analogy.analog_product_id:
        analog = products_by_id.get(analogy.analog_product_id) or db.get(
            Product, analogy.analog_product_id
        )
        if analog is None:
            warn.append(f"analogy {analogy.id}: analog product missing — ignored")
            return None
        rows = monthly.get(analog.id)
        if rows is None:
            rows = {
                (y, m): float(u or 0)
                for y, m, u in db.execute(
                    select(SalesMonthly.year, SalesMonthly.month, SalesMonthly.units).where(
                        SalesMonthly.product_id == analog.id
                    )
                )
            }
        series = _series_from_monthly_rows([(y, m, u) for (y, m), u in rows.items()], today)
        if not series.points:
            warn.append(
                f"analogy {analogy.id}: analog {analog.global_sku} has no sales history — ignored"
            )
            return None
        base = forecast_demand(series, rules.horizon, rules.forecast)
        return Forecast(
            monthly=base.monthly,
            method=METHOD_ANALOGY,
            baseline=base.baseline,
            confidence="low",
            n_history_months=0,
            low_data=True,
            uncertainty_pct=max(base.uncertainty_pct, 0.5),
            diverges_from_baseline=False,
            notes=[f"forecast by analogy with {analog.name or analog.global_sku}"]
            + ([analogy.rationale] if analogy.rationale else []),
        )
    if analogy.monthly_estimate and analogy.monthly_estimate > 0:
        return flat_forecast(
            analogy.monthly_estimate,
            rules.horizon,
            method=METHOD_ANALOGY,
            notes=["manual monthly estimate (no sales history yet)"],
        )
    return None


# ------------------------------------------------------------- upload paths
def build_bundle_from_workbook(
    db: Session, rules: OrderingRules, file_bytes: bytes
) -> SnapshotBundle:
    """Standalone inputs from the USA INV CHK workbook's SEA sheet: MON SALES,
    INV OH, per-row MTHS REQ (kept as the target override — the buyer curated
    it), and the six incoming-MOH columns."""
    import openpyxl

    _warnings.filterwarnings("ignore", module="openpyxl")
    bundle = SnapshotBundle(source="workbook")
    # read_only keeps an uploaded workbook from being materialized cell-by-cell
    # in memory; the cost is a forward-only sheet, so the loop below streams
    # rows and never touches ws.cell().
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    if "SEA" not in wb.sheetnames:
        bundle.warnings.append("workbook has no SEA sheet — nothing imported")
        wb.close()
        return bundle
    ws = wb["SEA"]
    by_sku = {
        p.global_sku: p
        for p in db.execute(
            select(Product).options(selectinload(Product.tags)).where(not_blacklisted())
        )
        .scalars()
        .all()
    }
    seen: set[str] = set()
    # NAME(1) SKU(2) CATEGORY(3) MON SALES(6) INV OH(7) MTHS REQ(10) and the six
    # INCOMING MOH columns 22..27 — 27 is the last column this loop reads.
    last_col = 27
    for raw in ws.iter_rows(min_row=2, max_col=last_col, values_only=True):
        def g(i: int, _row: tuple[Any, ...] = raw) -> Any:
            return _row[i - 1] if i - 1 < len(_row) else None

        sku = str(g(2) or "").strip()
        mon, oh, target = g(6), g(7), g(10)
        if not sku or sku in seen or not isinstance(mon, int | float) or mon <= 0:
            continue
        seen.add(sku)
        inc_moh = [v if isinstance(v, int | float) else 0.0 for v in (g(i) for i in range(22, 28))]
        oh = float(oh) if isinstance(oh, int | float) else 0.0
        product = by_sku.get(sku)
        if product is not None:
            pi = product_to_input(product)
            if isinstance(target, int | float):
                pi.target_moh_override = float(target)
        else:
            pi = ProductInput(
                global_sku=sku,
                name=str(g(1) or ""),
                category=str(g(3) or ""),
                target_moh_override=float(target) if isinstance(target, int | float) else None,
            )
            bundle.warnings.append(f"{sku}: not in the catalog — imported from workbook only")
        mon = float(mon)
        bundle.snapshots.append(
            SkuSnapshot(
                product=pi,
                on_hand=oh,
                avg_monthly_sales=mon,
                incoming_units_by_month=[m * mon for m in inc_moh][: rules.horizon],
                forecast=flat_forecast(
                    mon, rules.horizon,
                    notes=["workbook import: flat monthly average (no monthly history)"],
                ),
                units_sold=0.0,
                months_active=0,
            )
        )
    if not bundle.snapshots:
        bundle.warnings.append("no numeric SEA rows found in the workbook")
    wb.close()
    return bundle


def build_bundle_from_sales_csv(
    db: Session,
    rules: OrderingRules,
    file_bytes: bytes,
    today: date | None = None,
) -> SnapshotBundle:
    """Long-format sales CSV (`global_sku,year,month,units[,is_stockout]`)
    replacing the sales history; stock and incoming still come from the app
    snapshot. Unknown SKUs are warned about, never silently dropped."""
    today = today or utcnow().date()
    bundle = SnapshotBundle(source="csv")
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    field_names = {(f or "").strip().lower() for f in (reader.fieldnames or [])}
    required = {"global_sku", "year", "month", "units"}
    if not required.issubset(field_names):
        bundle.warnings.append(
            "CSV must have columns global_sku,year,month,units[,is_stockout]"
        )
        return bundle

    series_by_sku: dict[str, list[MonthPoint]] = {}
    bad_rows = 0
    for row in reader:
        clean = {(k or "").strip().lower(): (v or "").strip() for k, v in row.items()}
        try:
            pt = MonthPoint(
                year=int(clean["year"]),
                month=int(clean["month"]),
                units=float(clean["units"]),
                is_stockout=clean.get("is_stockout", "").lower() in ("1", "true", "yes"),
            )
        except (KeyError, ValueError):
            bad_rows += 1
            continue
        if not 1 <= pt.month <= 12:
            bad_rows += 1
            continue
        series_by_sku.setdefault(clean["global_sku"], []).append(pt)
    if bad_rows:
        bundle.warnings.append(f"{bad_rows} unparseable CSV row(s) skipped")

    by_sku = {
        p.global_sku: p
        for p in db.execute(
            select(Product).options(selectinload(Product.tags)).where(not_blacklisted())
        )
        .scalars()
        .all()
    }
    stock: dict[int, float] = {}
    for pid, qty in db.execute(
        select(StockLevel.product_id, StockLevel.qty).where(
            StockLevel.location_key.in_(ON_HAND_LOCATION_KEYS)
        )
    ):
        stock[pid] = stock.get(pid, 0.0) + float(qty or 0)
    incoming: dict[int, list[float]] = {}
    for pid, qty, expected in db.execute(
        select(IncomingMove.product_id, IncomingMove.qty, IncomingMove.expected_date)
    ):
        if pid is None:
            continue
        idx = _month_index_from_today(expected, today)
        if idx > rules.horizon:
            continue
        arr = incoming.setdefault(pid, [0.0] * rules.horizon)
        arr[idx - 1] += float(qty or 0)

    for sku, pts in series_by_sku.items():
        product = by_sku.get(sku)
        if product is None:
            bundle.warnings.append(f"{sku}: not in the catalog — skipped")
            continue
        pts.sort(key=lambda p: (p.year, p.month))
        series = MonthlySalesSeries(points=pts)
        forecast = forecast_demand(series, rules.horizon, rules.forecast)
        bundle.snapshots.append(
            SkuSnapshot(
                product=product_to_input(product),
                on_hand=stock.get(product.id, 0.0),
                avg_monthly_sales=forecast.baseline,
                incoming_units_by_month=incoming.get(product.id, [0.0] * rules.horizon),
                forecast=forecast,
                units_sold=sum(p.units for p in series.useable()),
                months_active=len(series.useable()),
            )
        )
    return bundle
