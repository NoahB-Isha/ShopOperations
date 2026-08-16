"""Importer for `IL City Coordinators.xlsx` — the real, messy roster.

Realities this handles (seen in the actual sheet):
  * `US` + `canada new` are current; `Canada (old)` and `Old` are legacy -> skipped.
  * Zones are numbers (1-4) in one column and coordinator first names (Lili,
    Mik, Ravi, Vivek) in another; Canada is its own zone.
  * Two conflicting "active" columns ("Active?" and "Is Active as of Jan
    2026?" with '?', 'NA', blanks). "Active?" wins; unresolvable rows import
    as INACTIVE and are flagged for admin follow-up rather than guessed.
  * Missing emails/phones -> flagged, never silently skipped.
  * Emails with stray spaces ("x@bellsouth. net") -> repaired by normalization.
  * Notes carry shared product sets ("Austin shares Products with San
    Antonio") -> both centers get one shared_product_group label.
  * Canada sheet has continuation rows (blank city = extra terminal/contact
    for the previous center) and multiple contacts per center.

Parse is pure (workbook -> ParsedCenter list + report); apply writes rows and
optionally creates orderer users for contacts that are reachable.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from ..auth.service import normalize_email, normalize_phone
from ..models import Center, CenterContact, Role, RoleAssignment, User, Zone, ZoneKind

LEGACY_SHEETS = {"Canada (old)", "Old"}

ZONE_COORDINATORS = {"1": "Lili", "2": "Mik", "3": "Ravi", "4": "Vivek"}
CANADA_ZONE = "Canada"


def _add_contact(center: ParsedCenter, contact: ParsedContact) -> None:
    """Attach a contact, merging by name instead of duplicating — continuation
    rows in the Canada sheet repeat the same person to hold a second terminal."""
    if not (contact.name or contact.email or contact.phone):
        return
    for existing in center.contacts:
        if contact.name and existing.name.lower() == contact.name.lower():
            existing.email = existing.email or contact.email
            existing.phone = existing.phone or contact.phone
            existing.role_note = existing.role_note or contact.role_note
            return
    center.contacts.append(contact)


def zone_name_for(code: str | None, coordinator: str | None) -> str | None:
    code = (code or "").strip().rstrip(".0") or None
    coord = (coordinator or "").strip().title() or None
    if coord and coord in ZONE_COORDINATORS.values():
        num = next(k for k, v in ZONE_COORDINATORS.items() if v == coord)
        return f"Zone {num} ({coord})"
    if code and code in ZONE_COORDINATORS:
        return f"Zone {code} ({ZONE_COORDINATORS[code]})"
    return None


@dataclass
class ParsedContact:
    name: str = ""
    email: str | None = None
    phone: str | None = None
    role_note: str = ""


@dataclass
class ParsedCenter:
    name: str
    city: str = ""
    state: str = ""
    region: str = ""
    country: str = "US"
    zone_name: str | None = None
    zone_code: str | None = None
    is_active: bool = False
    activity_raw: str = ""
    address: str = ""
    stripe_terminal_name: str = ""
    stripe_terminal_serial: str = ""
    notes: str = ""
    shared_product_group: str | None = None
    followup_reasons: list[str] = field(default_factory=list)
    contacts: list[ParsedContact] = field(default_factory=list)


@dataclass
class ImportReport:
    sheets_processed: list[str] = field(default_factory=list)
    sheets_skipped: list[str] = field(default_factory=list)
    centers_parsed: int = 0
    centers_created: int = 0
    centers_updated: int = 0
    zones_created: list[str] = field(default_factory=list)
    contacts: int = 0
    users_created: int = 0
    shared_groups: dict[str, list[str]] = field(default_factory=dict)
    followups: list[dict] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    applied: bool = False

    def to_dict(self) -> dict:
        return self.__dict__ | {"followup_count": len(self.followups)}


# ---------------------------------------------------------------- parsing
def _norm_header(h) -> str:
    return re.sub(r"[^a-z]", "", str(h or "").lower())


def _col(headers: list[str], *candidates: str) -> int | None:
    """Find a column by fuzzy header match (sheet headers have typos)."""
    normed = [_norm_header(h) for h in headers]
    for cand in candidates:
        c = _norm_header(cand)
        for i, h in enumerate(normed):
            if h.startswith(c) or c.startswith(h) and h:
                return i
    return None


def _cell(row: tuple, idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    v = row[idx]
    return str(v).strip() if v is not None else ""


def _parse_active(active_col: str, jan_col: str) -> tuple[bool, str, bool]:
    """(is_active, raw, ambiguous). 'Active?' wins; else 'Is Active as of Jan
    2026?'; anything unresolvable -> inactive + ambiguous flag."""
    a = active_col.strip().lower()
    j = jan_col.strip().lower()
    if a in ("yes", "y"):
        return True, active_col, False
    if a in ("no", "n"):
        return False, active_col, False
    if j in ("yes", "y"):
        return True, jan_col, False
    if j in ("no", "n"):
        return False, jan_col, False
    raw = active_col or jan_col or ""
    return False, raw, True


Sheet = tuple[str, list[str], list[tuple]]
"""(name, header row, data rows) — the shape both readers hand the parser."""


def _sheets_from_xlsx(path: Path) -> list[Sheet]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    out: list[Sheet] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(c.value or "") for c in ws[1]]
        out.append((sheet_name, headers, list(ws.iter_rows(min_row=2, values_only=True))))
    return out


def _sheets_from_csv(path: Path) -> list[Sheet]:
    """A CSV is one sheet. It therefore can't carry the workbook's per-sheet
    zone split, so the Zone column has to do that work — which it already does
    for the US sheets. The sheet is named after the file so the report reads
    the same either way."""
    import csv

    with path.open(newline="", encoding="utf-8-sig") as fh:
        rows = [tuple(r) for r in csv.reader(fh)]
    if not rows:
        return []
    return [(path.stem, [str(h or "") for h in rows[0]], rows[1:])]


def read_sheets(path: Path) -> list[Sheet]:
    """Rows from a .xlsx or a .csv. Anything else is refused by name rather
    than by a parser exception three frames down."""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return _sheets_from_xlsx(path)
    if suffix == ".csv":
        return _sheets_from_csv(path)
    raise ValueError(f"Unsupported roster file type '{suffix or path.name}' — use .xlsx or .csv.")


def parse_workbook(path: Path) -> tuple[list[ParsedCenter], ImportReport]:
    report = ImportReport()
    centers: dict[str, ParsedCenter] = {}

    for sheet_name, headers, data_rows in read_sheets(path):
        if sheet_name in LEGACY_SHEETS:
            report.sheets_skipped.append(sheet_name)
            continue
        is_canada = "canada" in sheet_name.lower()
        col = {
            "city": _col(headers, "City"),
            "state": _col(headers, "State"),
            "region": _col(headers, "Region"),
            "role": _col(headers, "City Coordinator Role"),
            "name": _col(headers, "Name"),
            "email": _col(headers, "Email"),
            "phone": _col(headers, "Phone"),
            "address": _col(headers, "Address"),
            "terminal_name": _col(headers, "Stripe Terminal Name"),
            "terminal_sn": _col(headers, "Stripe Terminal SN", "Stripe Serial number"),
            "zone": _col(headers, "Zone"),
            "active_jan": _col(headers, "Is Active as of Jan"),
            "notes": _col(headers, "Notes"),
            "zone_coord": _col(headers, "Zone Coordinator"),
            "active": _col(headers, "Active?"),
        }
        report.sheets_processed.append(sheet_name)
        current: ParsedCenter | None = None

        for row in data_rows:
            if not any(v not in (None, "") for v in row):
                continue
            city = _cell(row, col["city"])
            contact = ParsedContact(
                name=_cell(row, col["name"]),
                email=normalize_email(_cell(row, col["email"])),
                phone=normalize_phone(_cell(row, col["phone"])),
                role_note=_cell(row, col["role"]),
            )

            if not city:
                # continuation row: extra terminal and/or extra contact for the previous center
                if current is None:
                    report.warnings.append(f"{sheet_name}: orphan continuation row skipped")
                    continue
                tname = _cell(row, col["terminal_name"])
                tsn = _cell(row, col["terminal_sn"])
                if tname and not current.stripe_terminal_name:
                    current.stripe_terminal_name = tname
                elif tname:
                    current.stripe_terminal_name += f" / {tname}"
                if tsn and not current.stripe_terminal_serial:
                    current.stripe_terminal_serial = tsn
                elif tsn:
                    current.stripe_terminal_serial += f" / {tsn}"
                _add_contact(current, contact)
                continue

            key = city.lower()
            if key in centers:
                # same center listed again (second Canada contact) — merge
                current = centers[key]
                _add_contact(current, contact)
                continue

            active, raw, ambiguous = _parse_active(
                _cell(row, col["active"]), _cell(row, col["active_jan"])
            )
            if is_canada:
                # 'canada new' predates the active audit; explicit No is rare —
                # default the current sheet to active.
                explicit_no = raw.strip().lower() in ("no", "n")
                active, ambiguous = (not explicit_no), False

            parsed = ParsedCenter(
                name=city,
                city=city,
                state=_cell(row, col["state"]),
                region=_cell(row, col["region"]),
                country="CA" if is_canada else "US",
                zone_code=_cell(row, col["zone"]) or None,
                zone_name=CANADA_ZONE
                if is_canada
                else zone_name_for(_cell(row, col["zone"]), _cell(row, col["zone_coord"])),
                is_active=active,
                activity_raw=raw,
                address=_cell(row, col["address"]),
                stripe_terminal_name=_cell(row, col["terminal_name"]),
                stripe_terminal_serial=_cell(row, col["terminal_sn"]),
                notes=_cell(row, col["notes"]),
            )
            if ambiguous:
                parsed.followup_reasons.append("ambiguous_active")
            if parsed.zone_name is None:
                parsed.followup_reasons.append("no_zone")
            _add_contact(parsed, contact)
            if "temporary" in parsed.notes.lower():
                parsed.followup_reasons.append("temporary")
            centers[key] = parsed
            current = parsed

    result = list(centers.values())

    # contact reachability
    for c in result:
        if not any(ct.email or ct.phone for ct in c.contacts):
            c.followup_reasons.append("no_reachable_contact")
        elif any(not ct.email and not ct.phone for ct in c.contacts):
            c.followup_reasons.append("contact_missing_email_and_phone")
        else:
            for ct in c.contacts:
                if not ct.email:
                    c.followup_reasons.append("contact_missing_email")
                    break

    _assign_shared_groups(result, report)
    report.centers_parsed = len(result)
    report.contacts = sum(len(c.contacts) for c in result)
    report.followups = [
        {"center": c.name, "reasons": sorted(set(c.followup_reasons))}
        for c in result
        if c.followup_reasons
    ]
    return result, report


_SHARE_RE = re.compile(r"shares?\s+(products?|display set)\s+with\s+([A-Za-z .()-]+)", re.I)


def _assign_shared_groups(centers: list[ParsedCenter], report: ImportReport) -> None:
    def find(name: str) -> ParsedCenter | None:
        n = name.strip().lower().rstrip(".")
        for c in centers:
            cn = c.name.lower()
            if n == cn or n in cn or cn in n:
                return c
        return None

    for c in centers:
        m = _SHARE_RE.search(c.notes or "")
        if not m:
            continue
        other = find(m.group(2))
        if other is None:
            report.warnings.append(
                f"{c.name}: shared-set note mentions {m.group(2)!r} but no matching center"
            )
            continue
        label = "-".join(sorted([_slug(c.name), _slug(other.name)]))
        group = c.shared_product_group or other.shared_product_group or label
        c.shared_product_group = group
        other.shared_product_group = group
        report.shared_groups.setdefault(group, [])
        for name in (c.name, other.name):
            if name not in report.shared_groups[group]:
                report.shared_groups[group].append(name)


def _slug(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")


# ---------------------------------------------------------------- applying
def apply_import(
    db: Session,
    parsed: list[ParsedCenter],
    report: ImportReport,
    create_users: bool = True,
) -> ImportReport:
    zones: dict[str, Zone] = {z.name: z for z in db.scalars(select(Zone))}

    def zone_for(c: ParsedCenter) -> Zone | None:
        if not c.zone_name:
            return None
        z = zones.get(c.zone_name)
        if z is None:
            z = Zone(
                name=c.zone_name,
                kind=ZoneKind.FIELD.value,
                sheet_code=(c.zone_code or "").rstrip(".0") or None,
            )
            db.add(z)
            db.flush()
            zones[z.name] = z
            report.zones_created.append(z.name)
        return z

    for c in parsed:
        zone = zone_for(c)
        center = db.scalar(select(Center).where(func.lower(Center.name) == c.name.lower()))
        if center is None:
            center = Center(name=c.name)
            db.add(center)
            report.centers_created += 1
        else:
            report.centers_updated += 1
        center.zone_id = zone.id if zone else None
        center.city, center.state, center.region = c.city, c.state, c.region
        center.country = c.country
        center.is_active = c.is_active
        center.activity_raw = c.activity_raw
        center.address = c.address
        center.stripe_terminal_name = c.stripe_terminal_name
        center.stripe_terminal_serial = c.stripe_terminal_serial
        center.notes = c.notes
        center.shared_product_group = c.shared_product_group
        center.needs_followup = bool(c.followup_reasons)
        center.followup_reasons = sorted(set(c.followup_reasons))
        db.flush()

        existing_contacts = {
            (ct.name.lower(), ct.email or "", ct.phone or ""): ct for ct in center.contacts
        }
        for pc in c.contacts:
            key = (pc.name.lower(), pc.email or "", pc.phone or "")
            contact = existing_contacts.get(key)
            if contact is None:
                contact = CenterContact(center_id=center.id)
                db.add(contact)
            contact.name = pc.name
            contact.email = pc.email or ""
            contact.phone = pc.phone or ""
            contact.role_note = pc.role_note
            db.flush()

            if create_users and (pc.email or pc.phone) and center.is_active:
                user = _get_or_create_user(db, pc.email, pc.phone, pc.name)
                if user._created:  # type: ignore[attr-defined]
                    report.users_created += 1
                contact.user_id = user.id
                _ensure_role(db, user.id, Role.CENTER_ORDERER, center_id=center.id)

    db.commit()
    report.applied = True
    return report


def _get_or_create_user(db: Session, email: str | None, phone: str | None, name: str) -> User:
    user = None
    if email:
        user = db.scalar(select(User).where(func.lower(User.email) == email))
    if user is None and phone:
        user = db.scalar(select(User).where(User.phone == phone))
    if user is None:
        user = User(email=email, phone=phone, display_name=name or (email or phone or "Unknown"))
        db.add(user)
        db.flush()
        user._created = True  # type: ignore[attr-defined]
    else:
        user._created = False  # type: ignore[attr-defined]
        if phone and not user.phone:
            user.phone = phone
    return user


def _ensure_role(db: Session, user_id: int, role: Role, center_id: int | None = None,
                 zone_id: int | None = None) -> None:
    existing = db.scalar(
        select(RoleAssignment).where(
            RoleAssignment.user_id == user_id,
            RoleAssignment.role == role.value,
            RoleAssignment.center_id == center_id,
            RoleAssignment.zone_id == zone_id,
        )
    )
    if existing is None:
        db.add(RoleAssignment(user_id=user_id, role=role.value, center_id=center_id, zone_id=zone_id))


def run_import(db: Session, path: Path, apply: bool = True, create_users: bool = True) -> ImportReport:
    if not path.exists():
        raise FileNotFoundError(f"Coordinator workbook not found at {path}")
    parsed, report = parse_workbook(path)
    if apply:
        apply_import(db, parsed, report, create_users=create_users)
    return report
